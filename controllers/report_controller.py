import os

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    PageBreak
)

from reportlab.lib.styles import \
    getSampleStyleSheet

from openpyxl import Workbook

from models.utilisateur import Utilisateur
from models.session import SessionUtilisateur
from models.recommandation import Recommandation
from models.parcours import Parcours
from models.commande import Commande
from models.erreur import Erreur
from models.audit_log import AuditLog
from models.campagne import Campagne

from services.audit_service import \
    AuditService


class ReportController:

    EXPORT_FOLDER = "exports"

    @staticmethod
    def create_export_folder():

        if not os.path.exists(
                ReportController.EXPORT_FOLDER
        ):

            os.makedirs(
                ReportController.EXPORT_FOLDER
            )

    # ===================================
    # EXPORT PDF
    # ===================================

    @staticmethod
    def export_pdf():

        try:

            ReportController.create_export_folder()

            file_path = os.path.join(

                ReportController.EXPORT_FOLDER,

                "rapport_ccc.pdf"
            )

            doc = SimpleDocTemplate(
                file_path
            )

            styles = getSampleStyleSheet()

            elements = []

            elements.append(

                Paragraph(

                    "CCC Orientation System",

                    styles["Title"]
                )
            )

            elements.append(
                Spacer(1, 20)
            )

            # ==========================
            # STATISTIQUES
            # ==========================

            nb_utilisateurs = \
                Utilisateur.query.count()

            nb_sessions = \
                SessionUtilisateur.query.count()

            nb_recommandations = \
                Recommandation.query.count()

            nb_commandes = \
                Commande.query.count()

            nb_erreurs = \
                Erreur.query.count()

            nb_audits = \
                AuditLog.query.count()

            nb_campagnes = \
                Campagne.query.count()

            elements.append(

                Paragraph(
                    "Statistiques Générales",
                    styles["Heading1"]
                )
            )

            elements.append(

                Paragraph(
                    f"Utilisateurs : "
                    f"{nb_utilisateurs}",

                    styles["Normal"]
                )
            )

            elements.append(

                Paragraph(
                    f"Sessions : "
                    f"{nb_sessions}",

                    styles["Normal"]
                )
            )

            elements.append(

                Paragraph(
                    f"Recommandations : "
                    f"{nb_recommandations}",

                    styles["Normal"]
                )
            )

            elements.append(

                Paragraph(
                    f"Commandes : "
                    f"{nb_commandes}",

                    styles["Normal"]
                )
            )

            elements.append(

                Paragraph(
                    f"Erreurs : "
                    f"{nb_erreurs}",

                    styles["Normal"]
                )
            )

            elements.append(

                Paragraph(
                    f"Audits : "
                    f"{nb_audits}",

                    styles["Normal"]
                )
            )

            elements.append(

                Paragraph(
                    f"Campagnes : "
                    f"{nb_campagnes}",

                    styles["Normal"]
                )
            )

            elements.append(
                PageBreak()
            )

            # ==========================
            # UTILISATEURS
            # ==========================

            elements.append(

                Paragraph(
                    "Utilisateurs",
                    styles["Heading1"]
                )
            )

            for user in \
                    Utilisateur.query.all():

                elements.append(

                    Paragraph(

                        f"{user.nom} "
                        f"{user.prenom} | "
                        f"Age : {user.age} | "
                        f"Profil : "
                        f"{user.type_profil}",

                        styles["Normal"]
                    )
                )

            elements.append(
                PageBreak()
            )

            # ==========================
            # RECOMMANDATIONS
            # ==========================

            elements.append(

                Paragraph(
                    "Recommandations",
                    styles["Heading1"]
                )
            )

            for reco in \
                    Recommandation.query.all():

                parcours = \
                    Parcours.query.get(
                        reco.id_parcours
                    )

                nom_parcours = \
                    parcours.nom_parcours \
                    if parcours \
                    else "N/A"

                elements.append(

                    Paragraph(

                        f"Profil : "
                        f"{reco.profil_detecte} | "

                        f"Parcours : "
                        f"{nom_parcours} | "

                        f"Score : "
                        f"{reco.score}",

                        styles["Normal"]
                    )
                )

            doc.build(elements)

            AuditService.log(

                action="EXPORT",

                objet="PDF",

                resultat="SUCCES",

                details=file_path
            )

            return {

                "success": True,

                "file": file_path
            }

        except Exception as e:

            AuditService.log_error(
                str(e)
            )

            return {
                "success": False,
                "message": str(e)
            }

    # ===================================
    # EXPORT EXCEL
    # ===================================

    @staticmethod
    def export_excel():

        try:

            ReportController.create_export_folder()

            file_path = os.path.join(

                ReportController.EXPORT_FOLDER,

                "rapport_ccc.xlsx"
            )

            workbook = Workbook()

            # ==========================
            # UTILISATEURS
            # ==========================

            sheet_users = \
                workbook.active

            sheet_users.title = \
                "Utilisateurs"

            sheet_users.append([

                "ID",

                "Nom",

                "Prénom",

                "Age",

                "Profil",

                "Niveau"
            ])

            for user in \
                    Utilisateur.query.all():

                sheet_users.append([

                    user.id_utilisateur,

                    user.nom,

                    user.prenom,

                    user.age,

                    user.type_profil,

                    user.niveau_scolaire
                ])

            # ==========================
            # RECOMMANDATIONS
            # ==========================

            sheet_reco = \
                workbook.create_sheet(
                    "Recommandations"
                )

            sheet_reco.append([

                "ID",

                "Profil",

                "Score",

                "Parcours"
            ])

            for reco in \
                    Recommandation.query.all():

                parcours = \
                    Parcours.query.get(
                        reco.id_parcours
                    )

                sheet_reco.append([

                    reco.id_recommandation,

                    reco.profil_detecte,

                    reco.score,

                    parcours.nom_parcours
                    if parcours
                    else ""
                ])

            # ==========================
            # COMMANDES
            # ==========================

            sheet_cmd = \
                workbook.create_sheet(
                    "Commandes"
                )

            sheet_cmd.append([

                "ID",

                "Commande",

                "Résultat",

                "Valide"
            ])

            for cmd in \
                    Commande.query.all():

                sheet_cmd.append([

                    cmd.id_commande,

                    cmd.texte_commande,

                    cmd.resultat,

                    cmd.valide
                ])

            workbook.save(
                file_path
            )

            AuditService.log(

                action="EXPORT",

                objet="EXCEL",

                resultat="SUCCES",

                details=file_path
            )

            return {

                "success": True,

                "file": file_path
            }

        except Exception as e:

            AuditService.log_error(
                str(e)
            )

            return {

                "success": False,

                "message": str(e)
            }